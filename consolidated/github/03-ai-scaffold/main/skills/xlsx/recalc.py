import os
import sys
import subprocess
import json
import platform
from pathlib import Path
from openpyxl import load_workbook

def setup_libreoffice_macro():
    """
    Setup LibreOffice macro for recalculation if not already configured
    """
    macro_dir = os.environ.get('LIBREOFFICE_MACRO_DIR', 
        os.path.join(os.path.expanduser('~'), '.config/libreoffice/user/basic/Standard'))
    
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    
    if Path(macro_file).exists():
        with open(macro_file) as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not Path(macro_dir).exists():
        subprocess.run(['soffice', '--headless', '--terminate_after_init'], 
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)

    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'status': 'error', 'message': f'File {filename} does not exist'}
    
    abs_path = str(Path(filename).absolute())
    
    if not setup_libreoffice_macro():
        return {'status': 'error', 'message': 'Failed to setup LibreOffice macro'}
    
    cmd = [
        'soffice', '--headless', '--norestore',
        f'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]
    
    # Handle timeout command differences between Linux and macOS
    if platform.system() != 'Windows':
        timeout_cmd = f'timeout {timeout}' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # Check if gtimeout is available on macOS
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = f'gtimeout {timeout}'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass
        
        if timeout_cmd:
            cmd = [timeout_cmd] + cmd
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode != 0 and result.returncode != 124:  # 124 is timeout exit code
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'status': 'error', 'message': 'LibreOffice macro not configured properly'}
        else:
            return {'status': 'error', 'message': error_msg}
    
    try:
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check ALL rows and columns - no limits
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
```

**